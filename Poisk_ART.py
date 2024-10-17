import os
import pandas as pd
from datetime import datetime
import subprocess
import pytz  # Импортируем pytz для работы с временными зонами
from openpyxl import load_workbook
from openpyxl.styles import Alignment

# Функция для обновления файлов Excel из репозитория
def update_excel_files():
    try:
        subprocess.run(["git", "pull"], check=True)
        print("Файлы Excel обновлены из репозитория.")
    except subprocess.CalledProcessError:
        print("Ошибка при обновлении файлов Excel.")

# Обновляем данные перед началом работы
update_excel_files()

# Определяем путь к папке downloads
downloads_dir = "downloads"
result_file_name = "duplicate_items_across_files.xlsx"
output_file = os.path.join(downloads_dir, result_file_name)

# Считываем все Excel-файлы из папки downloads, исключая файл с результатами
excel_files = [f for f in os.listdir(downloads_dir) if f.endswith('.xlsx') and f != result_file_name]

# Создаем итоговый DataFrame для хранения объединенных данных
combined_df = pd.DataFrame()

# Считываем файлы и очищаем дубликаты
for excel_file in excel_files:
    file_path = os.path.join(downloads_dir, excel_file)

    try:
        df = pd.read_excel(file_path)
    except Exception as e:
        print(f"Ошибка при чтении файла {excel_file}: {e}")
        continue  # Пропускаем файл в случае ошибки

    # Проверяем наличие необходимых колонок
    required_columns = {'Артикул', 'Ссылка', 'Название', 'Цена'}
    if required_columns.issubset(df.columns):
        # Приводим колонку "Артикул" к строковому типу
        df['Артикул'] = df['Артикул'].astype(str)

        # Удаляем дубликаты по колонке "Артикул"
        df_unique = df.drop_duplicates(subset=['Артикул']).copy()

        # Переименовываем столбцы
        df_unique.rename(columns={
            'Название': f"Название_из_{os.path.splitext(excel_file)[0]}",
            'Ссылка': f"Ссылка_из_{os.path.splitext(excel_file)[0]}",
            'Цена': f"Цена_из_{os.path.splitext(excel_file)[0]}"
        }, inplace=True)

        # Добавляем данные в общий DataFrame
        if combined_df.empty:
            combined_df = df_unique[['Артикул', f"Название_из_{os.path.splitext(excel_file)[0]}",
                                      f"Ссылка_из_{os.path.splitext(excel_file)[0]}",
                                      f"Цена_из_{os.path.splitext(excel_file)[0]}"]]
        else:
            combined_df = pd.merge(combined_df,
                                   df_unique[['Артикул', f"Название_из_{os.path.splitext(excel_file)[0]}",
                                               f"Ссылка_из_{os.path.splitext(excel_file)[0]}",
                                               f"Цена_из_{os.path.splitext(excel_file)[0]}"]],
                                   on='Артикул', how='outer')

# Убираем артикулы, которые встречаются только в одном файле (оставляем те, которые есть хотя бы в двух)
non_empty_columns = combined_df.filter(like="Название_из_").notna().sum(axis=1)
duplicate_items = combined_df[non_empty_columns > 1]  # Оставляем только те строки, где артикулы встречаются хотя бы в двух файлах

# Добавляем текущую дату и время в отдельный столбец (с учетом временной зоны)
moscow_tz = pytz.timezone('Europe/Moscow')
duplicate_items.insert(0, 'Дата и время', datetime.now(moscow_tz).strftime('%Y-%m-%d %H:%M:%S'))

# Проверяем, существует ли файл с результатами
if os.path.exists(output_file):
    # Если файл существует, читаем его и добавляем новые данные
    old_data = pd.read_excel(output_file)
    final_data = pd.concat([old_data, duplicate_items], ignore_index=True)
else:
    # Если файла нет, создаем новый
    final_data = duplicate_items

# Сохраняем результаты в Excel-файл
final_data.to_excel(output_file, index=False)

# Форматирование: скрываем длинные ссылки и автоширина столбцов
wb = load_workbook(output_file)
ws = wb.active

# Устанавливаем перенос текста для всех ячеек
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(wrap_text=True)

# Автоширина для всех столбцов
for column_cells in ws.columns:
    length = max(len(str(cell.value)) for cell in column_cells)
    ws.column_dimensions[column_cells[0].column_letter].width = length + 2

# Сохраняем изменения в файл
wb.save(output_file)

print(f"Одинаковые товары найдены. Результаты сохранены в {output_file}")
